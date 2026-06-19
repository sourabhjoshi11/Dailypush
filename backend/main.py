from fastapi import FastAPI, HTTPException, Depends, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, StreamingResponse
import httpx
import os
from dotenv import load_dotenv
from supabase import create_client
from jose import jwt, JWTError
from datetime import datetime, timedelta, date
from pydantic import BaseModel
from typing import Optional
import base64
import io
import re
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import json

load_dotenv()

app = FastAPI(title="DailyPush API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://dailypush-seven.vercel.app",
        "http://localhost:5173",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Supabase
supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

# Config
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")
REDIRECT_URI = os.getenv("REDIRECT_URI")
FRONTEND_URL = os.getenv("FRONTEND_URL")
SECRET_KEY = os.getenv("SECRET_KEY", "secret")
GOOGLE_SCOPES = [
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
    "https://www.googleapis.com/auth/userinfo.profile",
]
N8N_WEBHOOK_URL = os.getenv("N8N_WEBHOOK_URL")
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")

# ── JWT helpers ──────────────────────────────────────────────
def create_jwt(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.utcnow() + timedelta(days=7),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")

def decode_jwt(token: str) -> dict:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    return decode_jwt(auth.split(" ")[1])

# ── Auth routes ───────────────────────────────────────────────
@app.get("/auth/google")
def google_login():
    scopes = "%20".join(GOOGLE_SCOPES)
    url = (
        f"https://accounts.google.com/o/oauth2/v2/auth"
        f"?client_id={GOOGLE_CLIENT_ID}"
        f"&redirect_uri={REDIRECT_URI}"
        f"&response_type=code"
        f"&scope={scopes}"
        f"&access_type=offline"
        f"&prompt=consent"
    )
    return RedirectResponse(url)

@app.get("/auth/callback")
async def google_callback(code: str):
    # Exchange code for tokens
    async with httpx.AsyncClient() as client:
        token_res = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code": code,
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "redirect_uri": REDIRECT_URI,
                "grant_type": "authorization_code",
            },
        )
    tokens = token_res.json()
    if "error" in tokens:
        raise HTTPException(status_code=400, detail=tokens["error"])

    access_token = tokens["access_token"]
    refresh_token = tokens.get("refresh_token", "")

    # Get user info
    async with httpx.AsyncClient() as client:
        user_res = await client.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {access_token}"},
        )
    user_info = user_res.json()
    email = user_info["email"]
    name = user_info.get("name", "")
    picture = user_info.get("picture", "")

    # Upsert user in Supabase
    existing = supabase.table("users").select("id").eq("email", email).execute()
    if existing.data:
        user_id = existing.data[0]["id"]
        supabase.table("users").update({
            "access_token": access_token,
            "refresh_token": refresh_token,
            "name": name,
            "picture": picture,
        }).eq("id", user_id).execute()
    else:
        res = supabase.table("users").insert({
            "email": email,
            "name": name,
            "picture": picture,
            "access_token": access_token,
            "refresh_token": refresh_token,
        }).execute()
        user_id = res.data[0]["id"]

    # Create JWT and redirect to frontend
    jwt_token = create_jwt(user_id, email)
    return RedirectResponse(f"{FRONTEND_URL}/auth?token={jwt_token}")

@app.get("/me")
def get_me(user=Depends(get_current_user)):
    res = supabase.table("users").select("id,email,name,picture").eq("id", user["sub"]).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="User not found")
    return res.data[0]

# ── Email send ────────────────────────────────────────────────
class SendEmailRequest(BaseModel):
    to: list[str]
    cc: Optional[list[str]] = []
    bcc: Optional[list[str]] = []
    subject: str
    body: str
    template_name: Optional[str] = ""

class RefineRequest(BaseModel):
    prompt: str

@app.post("/refine")
async def refine_text(payload: RefineRequest, user=Depends(get_current_user)):
    if not GROQ_API_KEY:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")
    async with httpx.AsyncClient() as client:
        res = await client.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
            json={
                "model": GROQ_MODEL,
                "messages": [
                    {"role": "system", "content": "You are a professional writing assistant. Refine and polish work updates. Be concise, clear, and professional. Return only plain text — no markdown, no bullet symbols, no ** bold **, no # headers. Always preserve field labels in the output formatted as Label: value on separate lines."},
                    {"role": "user", "content": payload.prompt}
                ],
                "max_tokens": 600,
            },
            timeout=20,
        )
    if not res.is_success:
        raise HTTPException(status_code=502, detail="Groq API error")
    data = res.json()
    raw = data["choices"][0]["message"]["content"].strip()
    clean = raw.replace("**", "").replace("__", "")
    return {"refined": clean}


@app.post("/transcribe")
async def transcribe_audio(file: UploadFile = File(...), user=Depends(get_current_user)):
    if not GROQ_API_KEY:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")

    try:
        async with httpx.AsyncClient() as client:
            file_bytes = await file.read()
            files = {
                "file": (file.filename, file_bytes, file.content_type or "application/octet-stream")
            }
            data = {
                "model": "whisper-large-v3",
                "language": "en",
                "prompt": "This is a professional work update by [USER]. It may include names, project names, and technical terms like MOM, blockers, dependencies.",
                "temperature": "0.2",
                "response_format": "json",
               
            }
            res = await client.post(
                "https://api.groq.com/openai/v1/audio/transcriptions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                data=data,
                files=files,
                timeout=60,
            )
        if res.status_code >= 400:
            error_detail = res.text
            raise HTTPException(status_code=502, detail=f"Groq API error: {error_detail}")
        data = res.json()
        text = data.get("text", "")
        return {"text": text}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))


def extract_task_text(body: str) -> Optional[str]:
    label_patterns = [
        r"Completed Tasks",
        r"Done",
        r"Accomplishments",
        r"Work Completed",
        r"What I accomplished",
    ]
    for label in label_patterns:
        pattern = re.compile(rf"(?:{label})\s*:\s*(.*?)(?=\n[A-Z][a-zA-Z /]*:|\Z)", re.DOTALL | re.IGNORECASE)
        match = pattern.search(body)
        if match:
            extracted = match.group(1).strip()
            return extracted if extracted else None
    return None


def detect_timesheet_structure(ws):
    best_date_row = None
    best_date_cells = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
        date_cells = {}
        for cell in row:
            if isinstance(cell.value, (datetime, date)):
                date_cells[cell.column - 1] = cell.value.date() if isinstance(cell.value, datetime) else cell.value
        if len(date_cells) > len(best_date_cells):
            best_date_cells = date_cells
            best_date_row = row_idx

    if not best_date_cells:
        return None, None, None

    ATTENDANCE_CODES = {"p", "present", "leave", "absent", "h", "public holiday", "week off", "weekend", "half day", "comp off"}
    task_row = None
    for row_idx in range(best_date_row + 1, min(best_date_row + 16, ws.max_row) + 1):
        cell_a_value = ws.cell(row=row_idx, column=1).value
        cell_b_value = ws.cell(row=row_idx, column=2).value
        if cell_a_value and str(cell_a_value).strip().lower() in ATTENDANCE_CODES:
            task_row = row_idx - 1
            break
        if cell_b_value and str(cell_b_value).strip().lower() in ATTENDANCE_CODES:
            task_row = row_idx - 1
            break

    if task_row is None:
        best_average_length = 0
        for row_idx in range(best_date_row + 1, best_date_row + 16):
            values = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            lengths = []
            for col_idx in best_date_cells.keys():
                cell_value = values[col_idx] if col_idx < len(values) else None
                if isinstance(cell_value, str) and cell_value.strip():
                    lengths.append(len(cell_value.strip()))
            if not lengths:
                continue
            average_length = sum(lengths) / len(lengths)
            if average_length >= 20 and average_length > best_average_length:
                best_average_length = average_length
                task_row = row_idx

    return best_date_row, best_date_cells, task_row


@app.post("/send")
async def send_email(payload: SendEmailRequest, user=Depends(get_current_user)):
    # Get user info from Supabase
    res = supabase.table("users").select("id,email,name").eq("id", user["sub"]).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="User not found")
    u = res.data[0]

    if not N8N_WEBHOOK_URL:
        raise HTTPException(status_code=500, detail="N8N_WEBHOOK_URL not configured")

    # Forward to n8n webhook with sender info
    webhook_payload = {
        "from": u["email"],
        "from_name": u["name"],
        "to": payload.to,
        "cc": payload.cc or [],
        "bcc": payload.bcc or [],
        "subject": payload.subject,
        "body": payload.body,
        "template_name": payload.template_name,
    }

    try:
        async with httpx.AsyncClient() as client:
            res_n8n = await client.post(N8N_WEBHOOK_URL, json=webhook_payload, timeout=15)
        if res_n8n.status_code >= 400:
            raise Exception(f"n8n responded with {res_n8n.status_code}")

        # Save to sent_emails history
        supabase.table("sent_emails").insert({
            "user_id": u["id"],
            "to_emails": payload.to,
            "cc_emails": payload.cc or [],
            "bcc_emails": payload.bcc or [],
            "subject": payload.subject,
            "body": payload.body,
            "template_name": payload.template_name,
            "status": "sent",
        }).execute()

        try:
            extracted = extract_task_text(payload.body)
            if extracted:
                entry_date = datetime.utcnow().date().isoformat()
                supabase.table("timesheet_entries").upsert({
                    "user_id": u["id"],
                    "entry_date": entry_date,
                    "task_text": extracted,
                    "updated_at": datetime.utcnow().isoformat(),
                }, on_conflict="user_id,entry_date").execute()
        except Exception as e:
            print(f"Timesheet incremental sync failed: {e}")

        return {"success": True, "message": "Email sent via n8n"}

    except Exception as e:
        supabase.table("sent_emails").insert({
            "user_id": u["id"],
            "to_emails": payload.to,
            "subject": payload.subject,
            "body": payload.body,
            "template_name": payload.template_name,
            "status": "failed",
        }).execute()
        raise HTTPException(status_code=500, detail=str(e))

# ── Sent emails history ───────────────────────────────────────
@app.get("/emails/history")
def get_history(user=Depends(get_current_user)):
    cutoff = (datetime.utcnow() - timedelta(days=7)).isoformat()
    res = supabase.table("sent_emails") \
        .select("*") \
        .eq("user_id", user["sub"]) \
        .gte("created_at", cutoff) \
        .order("created_at", desc=True) \
        .execute()
    return res.data

@app.post("/timesheet/upload")
async def upload_timesheet(file: UploadFile = File(...), user=Depends(get_current_user)):
    try:
        file_bytes = await file.read()
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
        except InvalidFileException:
            raise HTTPException(status_code=422, detail="Invalid Excel file. Please upload a valid .xlsx template.")
        ws = wb.active

        date_header_row, date_cells, task_row = detect_timesheet_structure(ws)
        if not date_cells or not task_row:
            raise HTTPException(status_code=422, detail="Could not detect timesheet structure. Please ensure the file has a row of dates and a row for task descriptions.")

        supabase.table("timesheet_templates").upsert({
            "user_id": user["sub"],
            "file_data": base64.b64encode(file_bytes).decode("ascii"),
            "date_header_row": date_header_row,
            "task_row": task_row,
            "uploaded_at": datetime.utcnow().isoformat(),
        }, on_conflict="user_id").execute()

        filled = 0
        total_days = len(date_cells)
        for col_idx, date_value in date_cells.items():
            start_of_day = datetime(date_value.year, date_value.month, date_value.day)
            end_of_day = start_of_day + timedelta(days=1)
            res = supabase.table("sent_emails") \
                .select("body") \
                .eq("user_id", user["sub"]) \
                .gte("created_at", start_of_day.isoformat()) \
                .lt("created_at", end_of_day.isoformat()) \
                .order("created_at", desc=True) \
                .limit(1) \
                .execute()

            if res.data:
                body_text = res.data[0].get("body", "") or ""
                extracted = extract_task_text(body_text)
                if extracted:
                    supabase.table("timesheet_entries").upsert({
                        "user_id": user["sub"],
                        "entry_date": date_value.isoformat(),
                        "task_text": extracted,
                        "updated_at": datetime.utcnow().isoformat(),
                    }, on_conflict="user_id,entry_date").execute()
                    filled += 1

        return {"filled": filled, "total_days": total_days}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/timesheet/preview")
def preview_timesheet(user=Depends(get_current_user)):
    res = supabase.table("timesheet_entries") \
        .select("entry_date,task_text") \
        .eq("user_id", user["sub"]) \
        .order("entry_date", asc=True) \
        .execute()
    return res.data or []


@app.get("/timesheet/download")
def download_timesheet(user=Depends(get_current_user)):
    res = supabase.table("timesheet_templates").select("*").eq("user_id", user["sub"]).limit(1).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="No timesheet template uploaded yet")

    template = res.data[0]
    file_data = template.get("file_data")
    if file_data is None:
        raise HTTPException(status_code=404, detail="No timesheet template uploaded yet")

    if isinstance(file_data, memoryview):
        file_bytes = file_data.tobytes()
    elif isinstance(file_data, (bytes, bytearray)):
        file_bytes = bytes(file_data)
    elif isinstance(file_data, str):
        try:
            file_bytes = base64.b64decode(file_data)
        except Exception:
            file_bytes = file_data.encode("utf-8")
    else:
        raise HTTPException(status_code=500, detail="Stored timesheet template has unsupported file data format")

    try:
        wb = load_workbook(io.BytesIO(file_bytes))
    except InvalidFileException:
        raise HTTPException(status_code=500, detail="Stored timesheet template is invalid or corrupted")
    ws = wb.active
    date_header_row = template.get("date_header_row")
    task_row = template.get("task_row")
    if not date_header_row or not task_row:
        raise HTTPException(status_code=500, detail="Stored timesheet template is missing structure metadata")

    date_columns = {}
    header_values = list(ws.iter_rows(min_row=date_header_row, max_row=date_header_row, values_only=True))[0]
    for col_idx, cell_value in enumerate(header_values, start=1):
        if isinstance(cell_value, (datetime, date)):
            date_columns[col_idx] = cell_value.date() if isinstance(cell_value, datetime) else cell_value

    entries = supabase.table("timesheet_entries") \
        .select("entry_date,task_text") \
        .eq("user_id", user["sub"]) \
        .order("entry_date", asc=True) \
        .execute()

    for entry in entries.data or []:
        entry_date = entry.get("entry_date")
        if isinstance(entry_date, str):
            entry_date = date.fromisoformat(entry_date)
        for col_idx, date_value in date_columns.items():
            if entry_date == date_value:
                ws.cell(row=task_row, column=col_idx).value = entry.get("task_text")
                break

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    headers = {"Content-Disposition": "attachment; filename=\"Timesheet_Current.xlsx\""}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# ── Templates ─────────────────────────────────────────────────
class TemplateModel(BaseModel):
    name: str
    icon: Optional[str] = "📝"
    raw_text: Optional[str] = ""
    fields: list
    ai_prompt: str

@app.get("/templates")
def get_templates(user=Depends(get_current_user)):
    res = supabase.table("templates").select("*").eq("user_id", user["sub"]).order("created_at").execute()
    return res.data

@app.post("/templates")
def create_template(template: TemplateModel, user=Depends(get_current_user)):
    res = supabase.table("templates").insert({
        "user_id": user["sub"],
        "name": template.name,
        "icon": template.icon,
        "raw_text": template.raw_text,
        "fields": template.fields,
        "ai_prompt": template.ai_prompt,
    }).execute()
    return res.data[0]

@app.delete("/templates/{template_id}")
def delete_template(template_id: str, user=Depends(get_current_user)):
    supabase.table("templates").delete().eq("id", template_id).eq("user_id", user["sub"]).execute()
    return {"success": True}

@app.get("/health")
def health():
    return {"status": "ok"}